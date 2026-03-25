"""
Build files for Vevox "Import from a file" (Excel).

Vevox flow (you do this in the browser):
  1. Dashboard -> Session -> Add Content -> Import from a file
  2. Download their **sample template** (important: column names change over time)
  3. Run this script with --template pointing to that file
  4. Upload the generated *_filled.xlsx back into Vevox

PowerPoint:
  Open your Vevox PPT add-in, sign in, pick the **same session**. Imported polls
  appear there for you to insert on slides. This repo cannot click the add-in for you.

Plan note: bulk Excel import may require Vevox Pro / Institution — check your account.

CSV fallback (no openpyxl): writes vevox_poll_bank.csv for manual copy into the template.

Usage:
  python export_vevox_import.py
  python export_vevox_import.py --template "%USERPROFILE%\\Downloads\\Vevox_import_template.xlsx"
  python export_vevox_import.py --topic Cardiology --limit 20
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "ite_data.json"


def _clean(s: str) -> str:
    if not s:
        return ""
    s = re.sub(r"[\uFFFD\uF000-\uF8FF]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _norm_header(s: object) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _load_questions(topic: str | None, year: int | None, limit: int | None) -> list[dict]:
    rows = json.loads(DATA.read_text(encoding="utf-8"))
    out = []
    for q in rows:
        if topic and topic.lower() not in (q.get("topic") or "").lower():
            continue
        if year is not None and q.get("year") != year:
            continue
        out.append(q)
        if limit and len(out) >= limit:
            break
    return out


def write_csv(questions: list[dict], path: Path) -> None:
    """Generic columns; paste into Vevox template if import rejects raw CSV."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "Topic",
                "Year",
                "Q",
                "QuestionText",
                "A",
                "B",
                "C",
                "D",
                "E",
                "CorrectLetter",
            ]
        )
        for q in questions:
            opts = q.get("options") or {}
            letters = sorted(opts.keys())
            def cell(L: str) -> str:
                return _clean(opts.get(L, ""))
            w.writerow(
                [
                    _clean(q.get("topic") or ""),
                    q.get("year", ""),
                    q.get("number", ""),
                    _clean(q.get("stem") or ""),
                    cell("A"),
                    cell("B"),
                    cell("C"),
                    cell("D"),
                    cell("E"),
                    (q.get("answer") or "").strip().upper()[:1],
                ]
            )


def _match_col(headers: list[str], aliases: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if _norm_header(h) in aliases:
            return i
    for i, h in enumerate(headers):
        n = _norm_header(h)
        for a in aliases:
            if a in n or n in a:
                return i
    return None


def fill_vevox_template(template: Path, questions: list[dict], out: Path) -> None:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit(
            "Install openpyxl:  pip install openpyxl\n"
            "Then re-run with --template ..."
        ) from e

    shutil.copy2(template, out)
    wb = openpyxl.load_workbook(out)
    # Prefer a sheet that is not "explanatory" notes
    sheet_name = None
    for name in wb.sheetnames:
        if "note" in name.lower():
            continue
        sheet_name = name
        break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    col_q = _match_col(
        row1,
        {
            "question",
            "question text",
            "poll question",
            "your question",
            "stem",
        },
    )
    col_a = _match_col(row1, {"a", "option a", "choice a", "answer 1", "answer a"})
    col_b = _match_col(row1, {"b", "option b", "choice b", "answer 2", "answer b"})
    col_c = _match_col(row1, {"c", "option c", "choice c", "answer 3", "answer c"})
    col_d = _match_col(row1, {"d", "option d", "choice d", "answer 4", "answer d"})
    col_e = _match_col(row1, {"e", "option e", "choice e", "answer 5", "answer e"})
    col_type = _match_col(
        row1,
        {"type", "poll type", "content type", "question type"},
    )
    col_correct = _match_col(
        row1,
        {
            "correct",
            "correct answer",
            "answer",
            "right answer",
            "marked correct",
        },
    )

    if col_q is None:
        raise SystemExit(
            "Could not find a Question column in row 1 of the template.\n"
            f"Headers seen: {row1[:25]!r}...\n"
            "Download a fresh Vevox sample template and try again."
        )

    example_type = None
    if col_type is not None and ws.max_row >= 2:
        example_type = ws.cell(row=2, column=col_type + 1).value

    start_row = 2
    # Clear previous data rows (optional): only from start_row downward for used cols
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    r = start_row
    for q in questions:
        opts = q.get("options") or {}

        def set_col(idx: int | None, val: str) -> None:
            if idx is not None:
                ws.cell(row=r, column=idx + 1).value = val

        if col_type is not None:
            set_col(col_type, str(example_type) if example_type else "Multiple Choice")
        set_col(col_q, _clean(q.get("stem") or ""))
        letters = sorted(opts.keys())
        mapping = [
            (col_a, "A"),
            (col_b, "B"),
            (col_c, "C"),
            (col_d, "D"),
            (col_e, "E"),
        ]
        for idx, L in mapping:
            if idx is not None:
                set_col(idx, _clean(opts.get(L, "")))
        if col_correct is not None:
            set_col(col_correct, (q.get("answer") or "").strip().upper()[:1])
        r += 1

    wb.save(out)
    print(f"Filled template -> {out}  ({len(questions)} questions, sheet={sheet_name!r})")


def main() -> int:
    ap = argparse.ArgumentParser(description="Export ITE MCQs for Vevox Excel import")
    ap.add_argument("--topic", help="Filter topic (substring)")
    ap.add_argument("--year", type=int, help="Filter year")
    ap.add_argument("--limit", type=int, help="Max questions")
    ap.add_argument(
        "--template",
        type=Path,
        help="Path to Vevox downloaded sample .xlsx (writes *_filled.xlsx next to it)",
    )
    args = ap.parse_args()

    if not DATA.exists():
        print("Missing ite_data.json — run extract_ite.py first.")
        return 1

    qs = _load_questions(args.topic, args.year, args.limit)
    if not qs:
        print("No questions matched filters.")
        return 1

    csv_out = ROOT / "vevox_poll_bank.csv"
    write_csv(qs, csv_out)
    print(f"Wrote {len(qs)} rows -> {csv_out}")

    if args.template:
        if not args.template.is_file():
            print(f"Template not found: {args.template}")
            return 1
        out = args.template.with_name(args.template.stem + "_filled.xlsx")
        fill_vevox_template(args.template, qs, out)
    else:
        print("\nNext: download Vevox's Import sample template, then:")
        print(f'  python export_vevox_import.py --template "PATH\\to\\template.xlsx"')
        if args.topic or args.year or args.limit:
            print("  (same filters as above)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
